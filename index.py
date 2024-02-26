"""
openpyxl
"""
from openpyxl import load_workbook
import os 
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

nome_arquivo = "C:/Users/Tecnologia/Documents/GitHub/RPA/InserirDadosPintarCelulas.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

sheet_selecionada = planilha_aberta['Aluno']

dadosTabela = [
    ['Nome', 'Idade'],
    ['Benerice', 28],
    ['Caio', 32],
    ['Nicole', 34],
    ['Leonardo', 19],
    ['Amanda', 25]
]

for linhaPlanilha in dadosTabela:
    sheet_selecionada.append(linhaPlanilha)

corTitulo = PatternFill(start_color='00FFFF00',
                       end_color='00FFFF00',
                       fill_type='solid')

sheet_selecionada["A1"].fill = corTitulo
sheet_selecionada["B1"].fill = corTitulo

for linha in range(2, len(sheet_selecionada['A']) + 1):
    celulaColunaA = "A" + linha

planilha_aberta.save(filename=nome_arquivo)

os.startfile(nome_arquivo)