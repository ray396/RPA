"""
Quebrando arquivos e criando emails
"""
from openpyxl import load_workbook
import os

nome_arquivo = "/GitHub/RPA/Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

sheet_selecionada = planilha_aberta['Dados']

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

for linha in range(2, len(sheet_selecionada['A']) + 1):
    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:
        
        linhaSheetQuebra = len(sheet_selecionada2['A']) + 1
        ceculaColunaA = "A" + str(linhaSheetQuebra)
        ceculaColunaB = "B" + str(linhaSheetQuebra)
        ceculaColunaC = "C" + str(linhaSheetQuebra)

        sheet_selecionada2[ceculaColunaA] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2[ceculaColunaB] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2[ceculaColunaC] = sheet_selecionada['C%s' % linha].value

    else:
        sheet_resumo = planilha_aberta.create_sheet(title=nomeAtual)
        sheet_selecionada2 = planilha_aberta[nomeAtual]
        nomeNovo = sheet_selecionada['A%s' % linha].value

        sheet_selecionada2['A1'] = "Vendedor"
        sheet_selecionada2['B1'] = "Produtos"
        sheet_selecionada2['C1'] = "Vendas"

        sheet_selecionada2['A2'] = sheet_selecionada['A%s' % linha].value
        sheet_selecionada2['B2'] = sheet_selecionada['B%s' % linha].value
        sheet_selecionada2['C2'] = sheet_selecionada['C%s' % linha].value

planilha_aberta.save(filename=nome_arquivo)

os.startfile(nome_arquivo)



"""
Quebrando arquivos e criando emails
"""
from openpyxl import load_workbook
from openpyxl import Workbook
import os

nome_arquivo = "/GitHub/RPA/Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

sheet_selecionada = planilha_aberta['Dados']
criandoNovoArquivoExcel = Workbook()

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

for linha in range(2, len(sheet_selecionada['A']) + 1):
    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:
        
        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        ceculaColunaA = "A" + str(linhaSheetQuebra)
        ceculaColunaB = "B" + str(linhaSheetQuebra)
        ceculaColunaC = "C" + str(linhaSheetQuebra)

        selecionaSheetVendasNovaPlanilha[ceculaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[ceculaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[ceculaColunaC] = sheet_selecionada['C%s' % linha].value

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

    else:
        nomeNovo = sheet_selecionada['A%s' % linha].value

        nova_planilha = criandoNovoArquivoExcel.active

        nova_planilha.title = "Vendas"

        caminhoNovaPlanilha = "/GitHub/RPA/"+ sheet_selecionada['A%s' % linha].value +".xlsx"

        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']

        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"

        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value

        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)




"""
Quebrando arquivos e criando emails
"""
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell

nome_arquivo = "/GitHub/RPA/Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

sheet_selecionada = planilha_aberta['Dados']
criandoNovoArquivoExcel = Workbook()

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

corAzul = PatternFill(start_color='0033cccc',
                      end_color='0033cccc',
                      fill_type='solid')

corAmarelo = PatternFill(start_color='00ccffcc',
                      end_color='00ccffcc',
                      fill_type='solid')
bdFina = Side(border_style="thin", color="000000")
borda = Border(left=bdFina, right=bdFina, top=bdFina, bottom=bdFina)

for linha in range(2, len(sheet_selecionada['A']) + 1):
    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:
        
        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        ceculaColunaA = "A" + str(linhaSheetQuebra)
        ceculaColunaB = "B" + str(linhaSheetQuebra)
        ceculaColunaC = "C" + str(linhaSheetQuebra)

        selecionaSheetVendasNovaPlanilha[ceculaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[ceculaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[ceculaColunaC] = sheet_selecionada['C%s' % linha].value


        selecionaSheetVendasNovaPlanilha[ceculaColunaA].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha[ceculaColunaB].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha[ceculaColunaC].fill = corAmarelo

        selecionaSheetVendasNovaPlanilha.column_dimensions["A"].width = 18
        selecionaSheetVendasNovaPlanilha.column_dimensions["B"].width = 25
        selecionaSheetVendasNovaPlanilha.column_dimensions["C"].width = 15

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

    else:
        nomeNovo = sheet_selecionada['A%s' % linha].value

        nova_planilha = criandoNovoArquivoExcel.active

        nova_planilha.title = "Vendas"

        caminhoNovaPlanilha = "GitHub/RPA/"+ sheet_selecionada['A%s' % linha].value +".xlsx"

        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']

        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"

        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value

        selecionaSheetVendasNovaPlanilha['A1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['B1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['C1'].fill = corAzul
        selecionaSheetVendasNovaPlanilha['A1'].border = borda
        selecionaSheetVendasNovaPlanilha['B1'].border = borda
        selecionaSheetVendasNovaPlanilha['C1'].border = borda

        selecionaSheetVendasNovaPlanilha['A2'].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha['B2'].fill = corAmarelo
        selecionaSheetVendasNovaPlanilha['C2'].fill = corAmarelo

        selecionaSheetVendasNovaPlanilha.delete_rows(3, 100)

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)



"""
Quebrando arquivos e criando emails
"""
import win32com.client as win32

outlook = win32.Dispatch('outlook.application')

emailOutlook = outlook.CreateItem(0)

emailOutlook.To = "rayssa@gmail.com"
emailOutlook.Subject = "Meu primeiro email usando python e outlook"
emailOutlook.HTMLBody = """
<p>Boa noite Rayssa</p>
<p>Esse é apenas um email de teste</p>
"""

emailOutlook.save()


"""
Quebrando arquivos e criando emails
"""
import win32com.client as win32

outlook = win32.Dispatch('outlook.application')

emailOutlook = outlook.CreateItem(0)

email = "testando@gmail.com"
variavelNome = "Rayssa"

emailOutlook.To = email
emailOutlook.Subject = "Meu primeiro email usando python e outlook"
emailOutlook.HTMLBody = f"""
<p>Bom dia! <b>{variavelNome}</b></p>
<p><font color="red"><b><u>Esse é apenas um email de teste</b></u></font></p>
<p><a href="https://www.google.com/">Clique aqui</a></p>
"""

emailOutlook.save()