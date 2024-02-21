import pyautogui as posicaoAbreArquivos

posicaoAbreArquivos.hotkey('win','r')

posicaoAbreArquivos.typewrite('notepad')

posicaoAbreArquivos.press('enter')

posicaoAbreArquivos.sleep(3)

posicaoAbreArquivos.typewrite('entrei atraves de codigo')

posicaoAbreArquivos.sleep(2)

fecharJanelaNotepad = posicaoAbreArquivos.getActiveWindow()

posicaoAbreArquivos.sleep(2)

fecharJanelaNotepad.close()

posicaoAbreArquivos.sleep(2)

posicaoAbreArquivos.press('tab')

posicaoAbreArquivos.sleep(2)

posicaoAbreArquivos.press('enter')


"""
Criando menu de opções para abrir o excel, word ou notepad
"""

import pyautogui
import pyautogui as escolha_opcao

opcao = pyautogui.confirm('clique no botão desejado', buttons = ['Excel', 'Word', 'Notepad'])

if opcao == 'Excel':
    escolha_opcao.hotkey('win', 'r')
    escolha_opcao.sleep(2)
    escolha_opcao.typewrite('Excel')
    escolha_opcao.press('Enter')
    escolha_opcao.sleep(6)
    escolha_opcao.click(x=434, y=273)
    escolha_opcao.sleep(3)
    escolha_opcao.typewrite('Escolhi abrir o Excel')
elif opcao == 'Word':
    escolha_opcao.hotkey('win', 'r')
    escolha_opcao.sleep(2)
    escolha_opcao.typewrite('winword')
    escolha_opcao.press('Enter')
    escolha_opcao.sleep(6)
    escolha_opcao.click(x=449, y=309)
    escolha_opcao.sleep(3)
    escolha_opcao.typewrite('Escolhi abrir o Word')
elif opcao == 'Notepad':
    escolha_opcao.hotkey('win', 'r')
    escolha_opcao.sleep(2)
    escolha_opcao.typewrite('notepad')
    escolha_opcao.press('Enter')
    escolha_opcao.sleep(2)
    escolha_opcao.typewrite('Escolhi abrir o Notepad')


from selenium import webdriver

abrirNavegador = webdriver.Chrome()
abrirNavegador.get("https://www.google.com.br/")


"""
Extraindo Dólar e Euro da internet
"""
from selenium import webdriver as opcoes_selenium
from selenium.webdriver.common.keys import Keys
import pyautogui as tempoPausaComputador
import pyautogui as teclasAtalhos
import xlsxwriter
import os

from selenium.webdriver.common.by import By

meuNavegador = opcoes_selenium.Chrome()
meuNavegador.get("https://www.google.com.br/")
tempoPausaComputador.sleep(2)

meuNavegador.find_element(By.NAME, "q").send_keys("Dolar hoje")
tempoPausaComputador.sleep(2)

meuNavegador.find_element(By.NAME, "q").send_keys(Keys.RETURN)
tempoPausaComputador.sleep(4)

valorDolar = meuNavegador.find_elements(By.XPATH, '//*[@id="knowledge-currency__updatable-data-column"]/div[1]/div[2]/span[1]')[0].text
print(valorDolar)

tempoPausaComputador.sleep(2)

meuNavegador.find_element(By.NAME, "q").send_keys("")
tempoPausaComputador.sleep(2)

teclasAtalhos.press('tab')

tempoPausaComputador.sleep(4)

teclasAtalhos.press('enter')

tempoPausaComputador.sleep(2)

meuNavegador.find_element(By.NAME, "q").send_keys("Euro hoje")
tempoPausaComputador.sleep(2)

meuNavegador.find_element(By.NAME, "q").send_keys(Keys.RETURN)
tempoPausaComputador.sleep(4)

valorEuro = meuNavegador.find_elements(By.XPATH, '//*[@id="knowledge-currency__updatable-data-column"]/div[1]/div[2]/span[1]')[0].text
print(valorEuro)

tempoPausaComputador.sleep(6)

caminhoArquivo = "C:/Users/Tecnologia/Documents/GitHub/RPA/dolar_e_euro.xlsx"
planilhaCriada = xlsxwriter.Workbook(caminhoArquivo)
sheet1 = planilhaCriada.add_worksheet()

sheet1.write("A1", "Dolar")
sheet1.write("B1", "Euro")
sheet1.write("A2", valorDolar)
sheet1.write("B2", valorEuro)

planilhaCriada.close()

os.startfile(caminhoArquivo)

"""
Extraindo endereço pelos sites dos correios e salvando no excel
"""
from openpyxl import load_workbook
import os

nome_aquivo_cep = "C:/Users/Tecnologia/Documents/GitHub/RPA/PesquisaEndereco_2.xlsx"
planilhaDadosEndereco = load_workbook(nome_aquivo_cep)

sheet_selecionada = planilhaDadosEndereco["CEP"]

from selenium import webdriver as opcoesSelenium
from selenium.webdriver.common.keys import Keys
import pyautogui as tempoEspera
from selenium.webdriver.common.by import By

navegador = opcoesSelenium.Chrome()
navegador.get("https://buscacepinter.correios.com.br/app/endereco/index.php")
tempoEspera.sleep(2)

navegador.find_element(By.NAME, "endereco").send_keys("23548057")
tempoEspera.sleep(2)
navegador.find_element(By.NAME, "btn_pesquisar").click()
tempoEspera.sleep(2)

for linha in range(2, len(sheet_selecionada['A']) + 1):
    tempoEspera.sleep(2)
    navegador.find_element(By.ID, "btn_nbusca").click()
    tempoEspera.sleep(2)

    cepPesquisa = sheet_selecionada['A%s' % linha].value
    tempoEspera.sleep(2)
    navegador.find_element(By.NAME, "endereco").send_keys(cepPesquisa)
    tempoEspera.sleep(2)
    navegador.find_element(By.NAME, "btn_pesquisar").click()
    tempoEspera.sleep(2)

    rua = navegador.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[1]')[0].text
    print(rua)
    bairro = navegador.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[2]')[0].text
    print(bairro)
    cidade = navegador.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[3]')[0].text
    print(cidade)
    cep = navegador.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[4]')[0].text
    print(cep)


    sheet_Dados = planilhaDadosEndereco["Dados"]
    linhaPlanilha = len(sheet_Dados['A']) + 1
    
    colunaA = "A" + str(linhaPlanilha)
    colunaB = "B" + str(linhaPlanilha)
    colunaC = "C" + str(linhaPlanilha)
    colunaD = "D" + str(linhaPlanilha)

    sheet_Dados[colunaA] = rua
    sheet_Dados[colunaB] = bairro
    sheet_Dados[colunaC] = cidade
    sheet_Dados[colunaD] = cep

planilhaDadosEndereco.save(filename=nome_aquivo_cep)

os.startfile(nome_aquivo_cep)